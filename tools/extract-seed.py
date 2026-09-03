#!/usr/bin/env python3
"""인사자료 xlsx → 시드 JSON 추출.

사용:
    .venv/bin/python3 tools/extract-seed.py <인사자료.xlsx> > docs/private/seed.json

출력은 개인정보(이메일 포함)를 담으므로 반드시 docs/private/ (git 제외)에만 쓴다.
휴대전화·내선·사번·호실은 어떤 경우에도 추출하지 않는다.
"""
import sys, json, os
import openpyxl

# 역할 명단은 **실명이라 저장소에 두지 않는다** (docs/private/, git 제외).
# 이 파일의 맨 위 규칙("개인정보는 docs/private/에만")이 정작 이 스크립트 자신에게는
# 지켜지지 않고 있었다 — 담당자 12명 이름이 여기 박혀 공개 저장소에 올라가 있었다.
#
#   docs/private/roles.json
#   { "leads": ["..."], "coordinators": ["..."], "operators": ["..."] }
_ROLES_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                           'docs', 'private', 'roles.json')
try:
    with open(_ROLES_PATH, encoding='utf-8') as _f:
        _roles = json.load(_f)
except FileNotFoundError:
    print(f'경고: {_ROLES_PATH} 가 없어 역할을 비워 둡니다 '
          '(담당자·총괄·운영자가 지정되지 않습니다).', file=sys.stderr)
    _roles = {}

KNOWN_LEADS = set(_roles.get('leads', []))
KNOWN_COORDINATORS = set(_roles.get('coordinators', []))
KNOWN_OPERATORS = set(_roles.get('operators', []))

# onRoster 기본 제외 직책 (DM-04: "팀장(실장)은 제외")
LEADERSHIP_TITLES = {'실장', '단장', '본부장', '센터장', '원장', '부원장', '감사'}

# 조직도(영문) 기반 슬러그. 사용자 예시 스타일: 언더스코어 연결 영문명
SLUGS = {
    '임원실': 'Executive_Office',                       # 조직도에 영문명 없음 → 관례적 명칭
    '감사실': 'Office_of_Audit',
    '글로벌대외협력단': 'Global_Cooperation_Bureau',
    '기획경영본부': 'Planning_and_Management_Department',
    '기획조정실': 'Planning_and_Strategy_Division',
    '연구관리실': 'Research_Management_Division',
    '인사관리실': 'Human_Resources_Division',
    'AI홍보전략실': 'AI_and_Public_Relations_Division',
    '경영지원실': 'Management_Support_Division',
    '기후대기전략연구본부': 'Climate_and_Atmospheric_Strategy_Research_Group',
    '탄소중립에너지연구실': 'Carbon_Neutrality_and_Energy_Division',
    '녹색경제연구실': 'Green_Economy_Division',
    '대기환경연구실': 'Atmospheric_Environment_Division',
    '생활환경연구본부': 'Living_Environment_Research_Group',
    '순환경제연구실': 'Circular_Economy_Division',
    '환경보건연구실': 'Environmental_Health_Division',
    '환경사회연구실': 'Environmental_Society_Division',
    '국토환경연구본부': 'Land_and_Environment_Research_Group',
    '국토관리연구실': 'Spatial_Planning_and_Management_Division',
    '물관리연구실': 'Water_Management_Division',
    '자연환경연구실': 'Natural_Environment_Division',
    '환경평가본부': 'Environmental_Assessment_Group',
    '국토전략평가실': 'Land_Strategy_Assessment_Division',
    '사회기반평가실': 'Social_Infrastructure_Assessment_Division',
    '기후에너지평가실': 'Climate_and_Energy_Assessment_Division',
    '환경자원평가실': 'Environmental_Resources_Assessment_Division',
    '국가기후위기적응센터': 'Korea_Adaptation_Center_for_Climate_Change',
    '기후적응정책실': 'Climate_Adaptation_Policy_Division',
    '기후적응협력실': 'Climate_Adaptation_Cooperation_Division',
    '국가지속가능발전연구센터': 'National_Research_Center_for_Sustainable_Development',
}


def short_slug(name_en_slug: str, taken: set[str]) -> str:
    """영문 슬러그 → 소문자 두문자 별칭. 충돌 시 첫 단어 접두를 늘려 재시도."""
    words = [w for w in name_en_slug.split('_') if w.lower() not in ('and', 'of', 'for', 'the')]
    base = ''.join(w[0] for w in words).lower()
    if base == 'aprd':               # AI는 두 글자를 살린다: aiprd
        base = 'aiprd'
    cand, prefix = base, 1
    while cand in taken:
        prefix += 1
        cand = (words[0][:prefix] + ''.join(w[0] for w in words[1:])).lower()
    taken.add(cand)
    return cand


def main(path: str) -> None:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    divisions: dict[str, dict] = {}
    cur_up = cur_dept = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        no, up, dept, sabun, name, emp, grade, role, duty, *_rest = r[:9]
        email = r[12]
        if up:
            cur_up = str(up).strip()
        if dept:
            cur_dept = str(dept).strip()
        if not name or not cur_dept:
            continue
        name = str(name).strip()
        title = str(role).strip() if role else ''
        d = divisions.setdefault(cur_dept, {
            'nameKo': cur_dept,
            'parentKo': cur_up,
            'slug': SLUGS.get(cur_dept),
            'shortSlug': None,
            'members': [],
        })
        d['members'].append({
            'name': name,
            'email': (str(email).strip().lower() if email else None),
            'grade': grade, 'title': title, 'duty': duty,
            'divisionRole': 'lead' if name in KNOWN_LEADS else 'member',
            'isCoordinator': name in KNOWN_COORDINATORS,
            'isOperator': name in KNOWN_OPERATORS,
            # DM-04: 실장급 직책은 기본 제외, 그 외 전원 포함 (열어둔다)
            'onRoster': title not in LEADERSHIP_TITLES,
        })

    taken: set[str] = set()
    for d in divisions.values():
        if d['slug']:
            d['shortSlug'] = short_slug(d['slug'], taken)

    out = {
        'source': path.split('/')[-1],
        'totalPeople': sum(len(d['members']) for d in divisions.values()),
        'divisions': list(divisions.values()),
    }
    json.dump(out, sys.stdout, ensure_ascii=False, indent=2)


if __name__ == '__main__':
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    main(sys.argv[1])
